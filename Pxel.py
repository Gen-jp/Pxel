from pathlib import Path
import tkinter
from tkinter import filedialog
import openpyxl
from tkinter import messagebox
import tkinter.simpledialog as simpledialog

class Application(tkinter.Frame):

    def __init__(self, root):
        super().__init__(root,width=720, height = 480, borderwidth = 4, relief='groove')
        self.Input_flag = 0
        self.root = root
        self.pack()
        self.pack_propagate(0)
        self.create_widgets()

    def create_widgets(self):
      

        self.submit_btn = tkinter.Button(self, width=10, )
        self.submit_btn['text'] = '開く'
        self.submit_btn['command'] = self.open_data
        self.submit_btn.place(x=0,y=0)

        #メッセージのオブジェクト作成
        self.Path_Label = tkinter.Label(self,font = ('MS Gothic', 10), foreground = '#000000')
        self.Path_Label.pack()

     
         

    def Title_Widgets(self):
        self.cell_texts = []
        self.cell_texslabel = []
        self.Title_text_box_Str = ('A1', 'B1', 'C1', 'D1','E1', 'F1','G1', 'H1', 'I1', 'J1')
        self.Title_Label = tkinter.Label(self, text=u'タイトル:上から順にA1～J1',font = ('MS Gothic', 20), foreground = '#FF00FF')    
        #self.Title_Label.place( x = 10, y = 50 )
        self.Title_Label.pack(anchor = tkinter.S)
        for i in range(10):
            #テキストボックスラベル
            Text_boxlab = self.Path_Label = tkinter.Label(self,font = ('MS Gothic', 10), foreground = '#000000')
            self.Path_Label['text']  = self.Title_text_box_Str[i]
            Text_boxlab.place(x='210', y=str((i * 30) + 100))
            # テキストボックス
            text_box = tkinter.Entry(self)
            text_box['width'] = 40
            text_box.place(x='250', y=str((i * 30) + 100))

            self.cell_texts.append(text_box)
            self.cell_texslabel.append(Text_boxlab)

        self.Save_button = tkinter.Button(self, width=10)
        self.Save_button['text'] = 'セット'
        self.Save_button['command'] = self.save_Title
        self.Save_button.place(x=250,y=400)



    def open_data(self):
        fTyp = [('XLSX', '.xlsx'), ('CSV', '.csv')] 

        self.open_file_name = tkinter.filedialog.askopenfilename(filetypes=fTyp, initialdir=Path.cwd())
        if(self.open_file_name != ""):
            self.submit_btn['state'] = "disable"
            self.wb = openpyxl.load_workbook(self.open_file_name)
            self.ws = self.wb.worksheets[0]
            self.Path_Label['text'] =    self.open_file_name
            self.Title_Widgets()
        
        #wb = openpyxl.load_workbook(file_name)
        #ws = wb.worksheets[0]

    def delete_Title_Wigets(self):
        self.Title_Label.destroy()
        self.Save_button.destroy()
        for i in range(10):
            self.cell_texts[i].destroy()
            self.cell_texslabel[i].destroy()
    
    def delete_Input_Wigets(self):

        for i in range(10):
            self.cell_Input_positions[i].destroy()
            self.cell_Input_texts[i].destroy()
            self.cell_Input_lab[i].destroy()
            self.cell_Input_Value_lab[i].destroy()
        self.Inputset_btn.destroy()
        self.Inputset_btn.destroy()

        


    def Init_Input_Data(self):
         Fisrt_sell = [u'A', u'B', u'C', u'D', u'E', u'F', u'G', u'H', u'I', u'J', ]
         for i in range(10):
                self.cell_Input_texts[i].delete(0, tkinter.END)
                self.cell_Input_positions[i].delete(0, tkinter.END)
                self.cell_Input_positions[i].insert(tkinter.END, Fisrt_sell[i])

    def Set_Input_Data(self):
         Fisrt_sell2 = [u'A', u'B', u'C', u'D', u'E', u'F', u'G', u'H', u'I', u'J', ]
         for i in range(10):
            if(self.cell_Input_texts[i].get() != '' and self.cell_Input_positions[i].get != '' and self.cell_Input_positions[i].get() != Fisrt_sell2[i]):
                  text = self.cell_Input_texts[i].get()
                  Positon =  self.cell_Input_positions[i].get()
                  self.ws[Positon].value = text

         self.Init_Input_Data()

    def Content_Input_Widgets(self):
        #self.submit_btn['state'] = "normal"
        self.cell_Input_positions = []
        self.cell_Input_texts = []
        self.cell_Input_lab = []
        self.cell_Input_Value_lab = []
        self.delete_Title_Wigets()

        for i in range(10):
            # メッセージ出力
            cell_label = tkinter.Message(self)
            cell_label['text'] = 'セル'
            cell_label.place(x='201', y=str((i * 22)+100))

            # テキストボックス
            position_box = tkinter.Entry(self)
            position_box['width'] = 4
            position_box.place(x='240', y=str((i * 22)+100))

            # メッセージ出力
            text_label = tkinter.Message(self)
            text_label['text'] = '値 '
            text_label.place(x='300', y=str((i * 22)+100))

            # テキストボックス
            text_box = tkinter.Entry(self)
            text_box['width'] = 20
            text_box.place(x='350', y=str((i * 22)+100))
        
            self.cell_Input_positions.append(position_box)
            self.cell_Input_texts.append(text_box)
            self.cell_Input_lab.append(text_label)
            self.cell_Input_Value_lab.append(cell_label)
        
        self.Init_Input_Data()

        # セットボタン
        self.Inputset_btn = tkinter.Button(self)
        self.Inputset_btn['text'] = "セット"
        self.Inputset_btn['command'] = self.Set_Input_Data
        self.Inputset_btn.pack(side="bottom")

         # 保存ボタン
        self.Inputsave_btn = tkinter.Button(self)
        self.Inputsave_btn['text'] = "保存"
        self.Inputsave_btn['command'] = self.All_Save
        self.Inputsave_btn.pack(side="bottom")


    def save_Title(self):
        for i in range(10):
            if(self.cell_texts[i].get() != ''):
                  text = self.cell_texts[i].get()
                  Positon = self.Title_text_box_Str[i]
                  self.ws[Positon].value = text
        
        self.ws.freeze_panes = 'A2'
        self.Content_Input_Widgets()
            
   
        #self.wb.save(self.open_file_name)
        #messagebox.showinfo("タイトルセット完了", "")


    def All_Save(self):
        self.wb.save(self.open_file_name)
        messagebox.showinfo("保存完了", "保存に成功しました")
        self.delete_Input_Wigets()
        self.submit_btn['state'] = "normal"

root = tkinter.Tk()
app = Application(root = root)
app.mainloop()
